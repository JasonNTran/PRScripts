import itertools
import os
import math
import re
import argparse
from typing import Optional
from collections.abc import Generator
import openpyxl
import numpy as np
from PIL import Image, ImageDraw, ImageFont, ImageChops


# Settings. Change these to your liking
BG_PATH = './Template/black.png'
FRAME_PATH = './Template/frame_cropped.png'
AVATAR_PATH = './avatars'
SAVE_PATH = "./PR"
X_PAD = 25
Y_PAD = 30
MIN_EDGE_PADDING = 40
AVATAR_SIZE = 125
NUM_COLS_PER_SIDE = 2
WIDTH = 1920
HEIGHT = 1090

# Where each box in the frame starts and ends based on pixel percentage. Hardcoded for this template
RANK_POSITION = [(0.096969697, 0.023784902), (0.172727273, 0.119958635)]
TOTAL_POSITION = [(0.853787879, 0.883143744), (0.928787879, 0.979317477)]
TYPE_POSITION = [(0.785606061, 0.006204757), (0.995454545, 0.055842813)]



""" The sheet needs several columns to work properly. It needs:
- An anime column containing "Anime" (Ex. Anime, Anime Info, Anime Name)
- A song name column called one of the following: 'song info', 'songinfo', 'songartist', "song name", "songname"
- A total column "Total"
- A rank column "Rank"
- A column for each ranker with their name matching their avatar icon file name
"""

global_guesses_dict = {}
class Person:
    def __init__(self, name: str, index: int, image: Image.Image):
        self.full_name = name
        self.index = index
        self.avatar = image

    def __eq__(self, other):
        return self.full_name == other.full_name

    @property
    def print_name(self):
        return clean_name(self.full_name, AVATAR_SIZE + X_PAD * 2 - 5)


class FontStyles:
    @staticmethod
    def load_fonts():
        return {
            "song": ImageFont.truetype("Fonts/Montserrat-Regular.ttf", size=30),
            "anime": ImageFont.truetype("Fonts/Montserrat-Regular.ttf", size=30),
            "type": ImageFont.truetype("Fonts/Montserrat-Regular.ttf", size=24),
            "rank": ImageFont.truetype("Fonts/Montserrat-Regular.ttf", size=72.5),
            "total": ImageFont.truetype("Fonts/Montserrat-Regular.ttf", size=52),
            "name": ImageFont.truetype("Fonts/antipasto.regular.ttf", size=30),
            "score": ImageFont.truetype("Fonts/SEANSBU.ttf", size=36),
            "guess": ImageFont.truetype("Fonts/antipasto.regular.ttf", size=26)
        }


class PanelConfig:
    def __init__(self,
                 mode: str,
                 info_dict: dict,
                 people: list[Person],
                 save_path: str,
                 centered: bool,
                 single_sided: bool,
                 guesses_dict):
        self.mode = mode
        self.info_dict = info_dict
        self.base_path = save_path
        self.centered = centered
        self.single_sided = single_sided
        self.fonts = FontStyles.load_fonts()
        self.background = Image.open(BG_PATH)
        self.video_frame = Image.open(FRAME_PATH)
        self.guesses_dict = guesses_dict
        self.people = people
        self.avatars = len(people)
        self.avatar_positions = None

    @property
    def offset(self) -> tuple[int, int]:
        bg_w, bg_h = self.background.size
        vf_w, vf_h = self.video_frame.size
        if self.single_sided == "off":
            return tuple(np.subtract(self.background.size, self.video_frame.size) // 2)
        elif self.single_sided == "right":
            return (X_PAD, (bg_h - vf_h) // 2)
        else:
            return (bg_w - vf_w - (X_PAD), (bg_h - vf_h) // 2)

    @property
    def frame_pos(self) -> list[int, int, int, int]:
        offset = self.offset
        vf_w, vf_h = self.video_frame.size
        return [offset[0], offset[0] + vf_w, offset[1], offset[1] + vf_h]

    def calculate_avatar_pos(
        self,
        space: tuple[int, int, int, int],
        count: int,
        cols=NUM_COLS_PER_SIDE
    ) -> Generator[tuple[int, int]]:
        if count != 0:
            left, right, top, bottom = space
            num_total_rows = int(math.ceil(count / cols))
            num_full_rows = count // cols
            remainder = count % cols
            vertical_size = num_total_rows * \
                AVATAR_SIZE + Y_PAD * (num_total_rows - 1)
            if self.centered:
                y_edge_padding = (bottom - top - vertical_size) / 2
            else:
                y_edge_padding = MIN_EDGE_PADDING / 2 if vertical_size + \
                    MIN_EDGE_PADDING > (top - bottom) else (bottom - vertical_size) / 2
            xpos = np.tile(
                np.arange(cols) * (AVATAR_SIZE + X_PAD) + (left + X_PAD),
                num_full_rows
            )
            num_per_row = np.repeat(cols, num_full_rows)
            if remainder > 0:
                extra_padding = (
                    (right - left) - (remainder * AVATAR_SIZE + (remainder - 1) * X_PAD)) / 2
                last_row = np.arange(
                    remainder) * (X_PAD + AVATAR_SIZE) + (left + extra_padding)
                xpos = np.append(xpos, last_row)
                num_per_row = np.append(num_per_row, remainder)
            ypos = np.repeat(
                np.arange(num_total_rows) * (AVATAR_SIZE + Y_PAD) +
                y_edge_padding + top,
                num_per_row
            )
            for i, j in zip(xpos, ypos):
                yield (int(i), int(j))

    def get_avatar_positions(self, inside_box_count: int, is_single_sided) -> list[tuple[int, int]]:
        frame_pos = self.frame_pos
        width, height = self.background.size

        if (is_single_sided) == 'off':
            left_count = (self.avatars - inside_box_count) // 2
            right_count = self.avatars - inside_box_count - left_count
        elif is_single_sided == 'left':
            left_count = self.avatars - inside_box_count
            right_count = 0
        elif is_single_sided == 'right':
            left_count = 0
            right_count = self.avatars - inside_box_count

        avatar_positions_left = self.calculate_avatar_pos(
            space=(0, frame_pos[0], 0, height),
            count=left_count,
        )
        avatar_positions_right = self.calculate_avatar_pos(
            space=(frame_pos[1], width, 0, height),
            count=right_count,
        )
        avatar_positions_inside = self.calculate_avatar_pos(
            space=(frame_pos[0], frame_pos[1], frame_pos[3] -
                   (AVATAR_SIZE + Y_PAD + 10), frame_pos[3]),
            count=inside_box_count,
            cols=inside_box_count
        )
        self.avatar_positions = list(itertools.chain(
            avatar_positions_left,
            avatar_positions_right,
            avatar_positions_inside))


def create_dirs(sheet_name: str) -> str:
    base_filename = os.path.splitext(os.path.basename(sheet_name))[0]
    save_path = os.path.join(SAVE_PATH, base_filename, "panels")
    os.makedirs(save_path, exist_ok=True)
    return save_path


def get_columns(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[dict, list[Person]]:
    info_dict = {
        "anime_column": None,
        "song_column": None,
        "type_column": None,
        "rank_column": None,
        "total_column": None,
        "nominator_column": None,
        "id_column": None
    }
    people = []
    for index, cell in enumerate(sheet[1]):
        column = cell.value
        if column:
            if info_dict["anime_column"] is None and "anime" in column.lower():
                info_dict["anime_column"] = index
            elif info_dict["song_column"] is None and column.lower() in ['song info', 'songinfo', 'songartist', "song artist", "song name", "songname", "artist"]:
                info_dict["song_column"] = index
            elif info_dict["type_column"] is None and "type" in column.lower():
                info_dict["type_column"] = index
            elif info_dict["rank_column"] is None and "rank" in column.lower():
                info_dict["rank_column"] = index
            elif info_dict["total_column"] is None and "total" in column.lower():
                info_dict["total_column"] = index
            elif info_dict["nominator_column"] is None and "nominator" in column.lower():
                info_dict["nominator_column"] = index
            elif info_dict["id_column"] is None and "id" in column.lower():
                info_dict["id_column"] = index
            elif info_dict["total_column"]:
                try:
                    path = os.path.join(AVATAR_PATH, f"{column}.png")
                    avatar = Image.open(path).resize(
                        (AVATAR_SIZE, AVATAR_SIZE))
                    people.append(Person(column, index, avatar))
                except FileNotFoundError:
                    print(f"[WARNING] Could not find image for {column}")
                    img = Image.new(
                        'RGBA', (AVATAR_SIZE, AVATAR_SIZE), (0, 0, 0, 255))
                    people.append(Person(column, index, img))
    return (info_dict, people)

def create_guesses_dict(wrkbk, dartboard_index: int):
    sheet = wrkbk.worksheets[dartboard_index]
    people_index_dict = {}
    guesses_dict = {}
    id_column = None

    for index, cell in enumerate(sheet[1]):
        column = cell.value
        if column == None:
            break
        elif id_column is None and "id" in column.lower():
            id_column = index
        elif column.lower() not in ["id", "nominator", "song name", "artist"]:
            people_index_dict[index] = column

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        song_id = row[id_column]
        guesses_dict[song_id] = {}
        for person_index, name in people_index_dict.items():
            guesses_dict[song_id][name] = row[person_index]
    return guesses_dict

def create_template(template_details: PanelConfig) -> Image.Image:
    template_image = Image.new('RGBA', template_details.background.size)
    template_image.paste(template_details.background)
    template_image.paste(template_details.video_frame, template_details.offset)

    return template_image


def create_song_panel(
    row: tuple,
    template_details: PanelConfig,
    template_panel: Image.Image,
) -> None:
    info_dict = template_details.info_dict
    panel = template_panel.copy()
    song_info = {
        "song_name": row[info_dict["song_column"]],
        "anime_name": row[info_dict["anime_column"]],
        "song_type": row[info_dict["type_column"]] if info_dict["type_column"] else None,
        "rank": row[info_dict["rank_column"]],
        "total": row[info_dict["total_column"]]
    }
    print(f"[INFO] Creating panel for song {song_info['song_name']}")
    if info_dict["nominator_column"] is None:
        nominator = None
        print(
            f"[WARNING] No nominator for song {song_info['song_name']} from show {song_info['anime_name']}")
    else:
        nominator = row[info_dict["nominator_column"]]

    write_song_info(song_info, panel, template_details)
    write_user_info(row, int(row[info_dict["id_column"]]), panel, nominator, template_details)

    save_path = os.path.join(template_details.base_path,
                             f"panel_{int(song_info['rank'])}.png")
    panel.save(save_path)


def write_song_info(song_info: dict, panel: Image.Image, template_details: PanelConfig) -> None:
    offset = template_details.offset
    bg_w, bg_h = template_details.background.size
    vf_w, vf_h = template_details.video_frame.size

    song_name = str(song_info["song_name"])
    anime_name = str(song_info["anime_name"])
    rank = str(int(song_info["rank"]))

    val = song_info["total"]
    total = str(int(val) if val.is_integer() else f"{float(val):.1f}")

    song_font = template_details.fonts["song"]
    anime_font = template_details.fonts["anime"]
    total_font = template_details.fonts["total"]
    rank_font = template_details.fonts["rank"]

    song_name_length = song_font.getlength(song_name)
    while (song_name_length > WIDTH - MIN_EDGE_PADDING * 2):
        font_size = song_font.size - 1
        song_font = ImageFont.truetype(
            "Fonts/Montserrat-Regular.ttf", font_size)
        song_name_length = song_font.getlength(song_name)

    draw = ImageDraw.Draw(panel)
    draw.rectangle(
        ((offset), (offset[0] + vf_w, offset[1] + vf_h)), outline=(193, 193, 193), width=1)
    draw.text((vf_w / 2 + offset[0], offset[1] / 2), song_name, font=song_font,
              fill='white', stroke_width=1, stroke_fill='black', anchor="mm")
    draw.text((vf_w / 2 + offset[0], (bg_h + offset[1] + vf_h) / 2), anime_name, font=anime_font,
              fill='white', stroke_width=1, stroke_fill='black', anchor="mm")

    rank_box_center = (offset[0] + (RANK_POSITION[0][0] * vf_w + RANK_POSITION[1][0] * vf_w) / 2,
                       offset[1] + (RANK_POSITION[0][1] * vf_h + RANK_POSITION[1][1] * vf_h) / 2)
    total_box_center = (offset[0] + (TOTAL_POSITION[0][0] * vf_w + TOTAL_POSITION[1][0] * vf_w) / 2,
                        offset[1] + (TOTAL_POSITION[0][1] * vf_h + TOTAL_POSITION[1][1] * vf_h) / 2)
    type_box_center = (offset[0] + (TYPE_POSITION[0][0] * vf_w + TYPE_POSITION[1][0] * vf_w) / 2,
                       offset[1] + (TYPE_POSITION[0][1] * vf_h + TYPE_POSITION[1][1] * vf_h) / 2)

    draw.text((rank_box_center), rank,
              font=rank_font, fill=(244, 186, 23),
              stroke_width=1, stroke_fill='black', anchor="mm")
    draw.text((total_box_center), total,
              font=total_font, fill=(244, 186, 23),
              stroke_width=1, stroke_fill='black', anchor="mm")
    if song_info["song_type"]:
        song_type = song_info["song_type"]
        type_font = template_details.fonts["type"]
        draw.text((type_box_center), song_type,
                  font=type_font, fill=(244, 186, 23),
                  stroke_width=1, stroke_fill='black', anchor="mm")


def write_user_info(row: tuple, song_id: int, panel: Image.Image, nominator: str, template_details: PanelConfig) -> None:
    low_people, high_people = getLowHighPeople(
        row, template_details.people, nominator)
    draw = ImageDraw.Draw(panel)

    for person, (start_x, start_y) in zip(template_details.people, template_details.avatar_positions):
        value = float(row[person.index])
        score = int(value) if value.is_integer() else f"{float(value):.1f}"
        text_color = "white"

        if template_details.mode == 'dartboard':
            if score < 4:
                border_color = (253, 255, 114) if score == 1 else \
                               (229, 229, 229) if score == 2 else \
                               (186, 137, 95)
                glow = create_glow(10)
                draw_glow(panel, (start_x - 10, start_y - 10),
                          glow, border_color)
            panel.paste(person.avatar, (start_x, start_y))
            score_box_tl = (start_x, int(start_y + AVATAR_SIZE - (AVATAR_SIZE * 0.3)))
            score_box_br = (int(start_x + AVATAR_SIZE * 0.3), start_y + AVATAR_SIZE)

            draw.rectangle(
                (score_box_tl, score_box_br),
                fill=(0,0,0),
                outline=(255, 255, 255), width=1
            )
            draw.rectangle(
                ((start_x, start_y), (start_x + AVATAR_SIZE, start_y + AVATAR_SIZE)),
                outline=(193, 193, 193), width=1
            )

            low_color = (53, 182, 31, 255)
            high_color = (255, 0, 0, 255)
            nom_color = (0, 255, 255, 255)
            score_pos = ((score_box_tl[0] + score_box_br[0]) // 2, (score_box_tl[1] + score_box_br[1]) // 2)
            if(person.full_name.lower() != nominator.lower()):
                # print(global_guesses_dict)
                guess = global_guesses_dict[song_id][person.full_name]
                guess_color = (241, 61, 66) if guess != nominator else (47, 193, 87)
                # guess_cleaned = clean_name(guess, AVATAR_SIZE * .75)
                guess_pos = (int(start_x + AVATAR_SIZE * 0.3) + 10, int(start_y + AVATAR_SIZE - template_details.fonts["guess"].size * 0.1))
                draw.text(
                    (guess_pos), guess, font=template_details.fonts["guess"], fill="white",
                    stroke_width=2, stroke_fill=guess_color,
                    anchor='lm'
                )

        else:
            score_pos = (start_x + AVATAR_SIZE / 2, 
                start_y + AVATAR_SIZE - template_details.fonts["score"].size * 0.2)
            if template_details.mode == 'ranking':
                if score < 4:
                    border_color = (253, 255, 114) if score == 1 else \
                                (229, 229, 229) if score == 2 else \
                                (186, 137, 95)
                    glow = create_glow(10)
                    draw_glow(panel, (start_x - 10, start_y - 10),
                            glow, border_color)

                low_color = (53, 182, 31, 255)
                high_color = (255, 0, 0, 255)
                nom_color = (0, 255, 255, 255)

            elif template_details.mode == "scoring":
                low_color = (255, 0, 0, 255)
                high_color = (53, 182, 31, 255)
                nom_color = (0, 255, 255, 255)
            else:
                raise ValueError(f"Unsupported mode: {template_details.mode}")

            if person in low_people:
                text_color = low_color
            elif person in high_people:
                text_color = high_color
            elif nominator and person.full_name.lower() == nominator.lower():
                text_color = nom_color

            panel.paste(person.avatar, (start_x, start_y))
            draw.rectangle(
                ((start_x, start_y), (start_x + AVATAR_SIZE, start_y + AVATAR_SIZE)),
                outline=(193, 193, 193), width=1
            )
        draw.text(
            (start_x + AVATAR_SIZE / 2, start_y), person.print_name, font=template_details.fonts["name"], fill="white",
            stroke_width=2, stroke_fill='black',
            anchor="mm"
        )
        draw.text(
            score_pos,
            str(score), font=template_details.fonts["score"],
            fill=text_color, stroke_width=2, stroke_fill='black',
            anchor="mm"
        )


def create_glow(glow_width: int) -> Image.Image:
    width = AVATAR_SIZE + glow_width * 2

    def calc_alpha(x, y):
        d = np.abs(
            np.stack([(x - width / 2), (y - width / 2)])) - AVATAR_SIZE / 2
        l = np.linalg.norm(np.maximum(d, 0), axis=0)
        # m = np.minimum(np.max(d, axis = 0), 0)
        distance = l
        alpha = np.maximum(1.0 - 1.0 / glow_width * distance, 0)
        return (alpha * 255).astype(np.uint8)
    data = np.fromfunction(calc_alpha, shape=(width, width), dtype=np.uint8)
    im = Image.fromarray(data, "L")
    return im


def draw_glow(
    panel: Image.Image,
    pos: tuple[int, int],
    glow: Image.Image,
    glow_color="yellow",
) -> None:
    glow_base = Image.new(mode="L", size=panel.size)
    x, y = pos
    tmp = Image.new(mode="L", size=panel.size)
    tmp.paste(glow, (x, y))
    # We don't use Image.paste here because we want to just take the whiter value
    glow_base = ImageChops.lighter(glow_base, tmp)
    # in theory we could paste the glows with a mask on a separate layer with the boxes
    new_im = Image.composite(
        Image.new(mode="RGBA", size=panel.size, color=glow_color),
        panel,
        glow_base,
    )
    panel.paste(new_im)


def clean_name(name: str, max_length) -> str:
    name = re.sub(r'\d+', '', name).strip()
    font = ImageFont.truetype("Fonts/antipasto.regular.ttf", size=30)
    name_length = font.getlength(name)
    while (name_length > max_length ):
        decrease = max_length / name_length
        num_char = len(name)
        name = name[:int(num_char * decrease)]
        name_length = font.getlength(name)
    return name


def getLowHighPeople(
        row: tuple,
        people: list[Person],
        nominator:
        str) -> tuple[list[Person], list[Person]]:
    low_score = None
    high_score = None
    low_people = []
    high_people = []
    for person in people:
        if (person.full_name.lower() != nominator.lower()):
            index = person.index
            score = row[index]
            if low_score is None or score < low_score:
                low_score = score
            if high_score is None or score > high_score:
                high_score = score

    for person in people:
        if (person.full_name.lower() != nominator.lower()):
            index = person.index
            score = row[index]
            if score == low_score:
                low_people.append(person)
            elif score == high_score:
                high_people.append(person)

    return (low_people, high_people)


def adjust_frame(template_details: PanelConfig, type_column: Optional[int], single_sided: str) -> Image.Image:
    bg_w, _ = template_details.background.size
    _, vf_h = template_details.video_frame.size
    
    space_per_side = (NUM_COLS_PER_SIDE + 1) * X_PAD + (NUM_COLS_PER_SIDE * AVATAR_SIZE)
    padding = 0 if single_sided == "off" else X_PAD
    sides = 2 if single_sided == "off" else 1
    avatar_space = space_per_side * sides + padding
    
    new_width = bg_w - avatar_space
    new_frame = template_details.video_frame.resize((new_width, vf_h))
    if (type_column is None):
        rect_width = int(
            (TYPE_POSITION[1][0] - TYPE_POSITION[0][0]) * new_width * 1.3)
        rect_height = int(
            (TYPE_POSITION[1][1] - TYPE_POSITION[0][1]) * vf_h * 1.3)
        rect_pos = (new_width - rect_width), 0

        transparent = Image.new(
            'RGBA', (rect_width, rect_height), (255, 0, 0, 0))

        new_frame.paste(transparent, rect_pos)
    return new_frame


def main(args) -> None:
    global global_guesses_dict
    sheet_name = args.sheet
    inside_box_count = args.inside_box
    mode = args.mode
    is_centered = args.centered
    is_single_sided = args.single_sided
    dartboard = args.dartboard

    save_path = create_dirs(sheet_name)
    file_path = os.path.join(os.getcwd(), sheet_name)
    wrkbk = openpyxl.load_workbook(sheet_name, data_only=True)
    sheet = wrkbk.worksheets[0]
    
    indices_info, people = get_columns(sheet)
    global_guesses_dict = create_guesses_dict(wrkbk, dartboard) if dartboard else None
    guesses_dict = {}
    template_details = PanelConfig(
        mode, indices_info, people, save_path, is_centered, is_single_sided, guesses_dict)

    template_details.video_frame = adjust_frame(
        template_details, indices_info["type_column"], is_single_sided)

    template_details.get_avatar_positions(inside_box_count, is_single_sided)
    template_panel = create_template(template_details)

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if row is None or row[0] is None:
            print(f"[INFO] Hit none on row {index + 2}. Exiting")
            break
        create_song_panel(row, template_details, template_panel)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Generate video panels from spreadsheet.')
    parser.add_argument('sheet', type=str, help='Path to the Excel sheet')
    parser.add_argument("-i", '--inside_box', type=int, nargs='?',
                        default=0, help='Number of people inside the video box')
    parser.add_argument("-m", '--mode', type=str, choices=[
                        'scoring', 'ranking', 'dartboard'], default='ranking', help='Type of PR. Scoring, or ranking')
    parser.add_argument("--centered", action="store_true",
                        help="Option to center the avatar boxes vertically")
    parser.add_argument("-s", '--single_sided', type=str, choices=['off',
                        'left', 'right'], default='off', help='Single sided mode')
    parser.add_argument("-d", '--dartboard', type=int, default=None, help='Sheet index of guesses for dartboard')

    if args.mode == "dartboard" and args.dartboard is None:
        parser.error("--dartboard must be provided when mode is 'dartboard'")

    args = parser.parse_args()
    main(args)
