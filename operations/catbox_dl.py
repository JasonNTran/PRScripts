import os
import subprocess
import sqlite3
import requests
import openpyxl
import argparse
from urllib.parse import urlparse
import re
import sys
import time

DOC_STRING = """
Usage:
  catbox_dl.py <command> <sheet.xlsx>
"""
exclude_artists = 0
include_rank = 0


def normalizeTime(time_str):
    print(time_str)
    if time_str is None:
        return None
    parts = str(time_str).split(':')
    minutes, seconds, _ = parts
    return f"00:{int(minutes):02d}:{int(seconds):02d}"


def cleanup_song(song):
    if exclude_artists == 1:
        song = "".join(song.split("by")[:-1])
        if ' by' in song:
            song = "".join(song.split("by")[:-1])
        elif ' BY' in song:
            song = "".join(song.split("BY")[:-1])
        else:
            print(
                f'[WARN] Could not split artist from song for song /{song.encode("utf-8")}/')
    song = song.replace("'", "")
    song = song.replace("<", "-")
    song = song.replace(">", "-")
    song = re.sub(r'[:"/\\|?*]', ' ', song)
    song = song.strip()
    return song


def get_columns(sheet, isMp3):
    song_column = None
    link_column = None
    rank_column = None
    start_column = None
    end_column = None
    artist_column = None
    for index, cell in enumerate(sheet[1]):
        column = cell.value
        if column:
            if column.lower() in ['song name', 'song info', 'songinfo', 'songartist'] and song_column is None:
                song_column = index
            if "artist" in column.lower() and not "songartist" in column.lower() and artist_column is None:
                artist_column = index
            elif 'mp3' in column.lower() and isMp3:
                link_column = index
            elif column.lower() in ['songlink', 'song link', "link"]:
                link_column = index
            elif "rank" in column.lower():
                rank_column = index
            elif "start" in column.lower():
                start_column = index
            elif "end" in column.lower():
                end_column = index
    if link_column == None:
        link_column = song_column
    return (song_column, link_column, rank_column, start_column, end_column, artist_column)


def dl_song(hostname, link, file_name, isMp3, start_time, end_time):
    if hostname in ["www.youtube.com", "youtu.be", "music.youtube.com"]:
        if isMp3:
            file = subprocess.run([
                "yt-dlp",
                "--encoding", "utf-8",
                "--no-playlist",
                "-f", "bestaudio/best",
                "--extract-audio",
                "-o", file_name,
                link
            ], encoding='utf-8')
        else:
            out_path = f"{file_name}.mp4"
            file = subprocess.run([
                "yt-dlp",
                "--encoding", "utf-8",
                "--no-playlist",
                "-f", "bestvideo+bestaudio/22/18",
                "--merge-output-format", "mp4",
                "-o", out_path,
                link
            ], encoding='utf-8')

    elif hostname in ["files.catbox.moe", "openings.moe", "ladist1.catbox.video", "naedist.animemusicquiz.com", "nawdist.animemusicquiz.com", "eudist.animemusicquiz.com"]:
        print(link)
        headers = {
            'User-agent': 'Mozilla/5.0'
        }
        response = requests.get(link, headers=headers)
        extension = link.split(".")[-1]
        out_path = f"{file_name}.{extension}" if not isMp3 else f"{file_name}.mp3"
        with open(out_path, "wb") as file:
            file.write(response.content)
        result = subprocess.run(
            ['ffmpeg', '-i', out_path, "-c", "copy", "-metadata", 'title='])
        if result.returncode != 0:
            print(f"ffmpeg failed with code {result.returncode}")
            return
    else:
        print("Hostname not recognized", hostname, file_name)
        sys.stdout.flush()
    if start_time is not None and end_time is not None and os.path.exists(out_path):
        base, ext = os.path.splitext(out_path)
        tmp_path = base + ".tmp" + ext
        result = subprocess.run(
            ['ffmpeg',  '-i', out_path, '-ss', str(start_time), "-to", str(end_time), "-c", "copy", tmp_path])
        os.replace(tmp_path, out_path)


def dl_ranks_mp3(file_name, index):
    folder = os.path.splitext(file_name)[0]
    if not os.path.exists(folder):
        os.mkdir(folder)

    wrkbk = openpyxl.load_workbook(file_name)
    sheet = wrkbk.worksheets[index]
    song_column, link_column, rank_column, start_column, end_column, artist_column = get_columns(
        sheet, True)

    for row in sheet.iter_rows(min_row=2):
        if row[0].value is None:
            print(
                "[INFO] None detected. This could be an error or end of file. Exiting")
            return
        song_name = row[song_column].value
        try:
            link = row[link_column].hyperlink.target
        except:
            try:
                link = row[song_column].hyperlink.target
            except:
                link, song_name = song_name.split('by')[1::2]
        song_name = cleanup_song(song_name)
        start_time = normalizeTime(
            row[start_column].value) if start_column is not None else None
        end_time = normalizeTime(
            row[end_column].value) if end_column is not None else None
        if include_rank == 1:
            rank = (int)(row[rank_column].value)
            file_name = f"{folder}/{rank}-{song_name}"
        else:
            file_name = f"{folder}/{song_name}"
        host_name = urlparse(link).hostname
        dl_song(host_name, link, file_name, isMp3=True,
                start_time=start_time, end_time=end_time)


def dl_vids(file_name, index):
    folder = os.path.splitext(file_name)[0]
    if not os.path.exists(folder):
        os.mkdir(folder)

    wrkbk = openpyxl.load_workbook(file_name)
    sheet = wrkbk.worksheets[index]
    song_column, link_column, _, start_column, end_column, artist_column = get_columns(
        sheet, False)

    count = 2
    for row in sheet.iter_rows(min_row=2):
        if row[0].value is None:
            print(
                "[INFO] None detected. This could be an error or end of file. Exiting")
            return
        song_name = row[song_column].value
        try:
            link = row[link_column].hyperlink.target
        except:
            link, song_name = song_name.split('by')[1::2]
        song_name = cleanup_song(song_name)
        host_name = urlparse(link).hostname
        file_name = f"{folder}/{song_name}"
        start_time = normalizeTime(
            row[start_column].value) if start_column is not None else None
        end_time = normalizeTime(
            row[end_column].value) if end_column is not None else None
        dl_song(host_name, link, file_name, isMp3=False,
                start_time=start_time, end_time=end_time)
        count += 1


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("sheet", type=str, help="The sheet to read from")
    parser.add_argument("-a", "--exclude_artist")
    parser.add_argument("-r", "--include_rank")
    parser.add_argument("-m", '--mode', type=str, choices=[
                        'mp3', 'mp4'], default='mp3', help='Type to download. mp3 or mp4')
    parser.add_argument("-i", '--sheet_index', type=int,
                        default='0', help='Index of the sheet to read from')

    args = parser.parse_args()
    command = args.mode
    sheet = args.sheet
    index = args.sheet_index
    if args.exclude_artist:
        exclude_artists = int(args.exclude_artist)
    if args.include_rank:
        include_rank = int(args.include_rank)

    if command == 'mp4':
        dl_vids(sheet, index)
    elif command == 'mp3':
        dl_ranks_mp3(sheet, index)
    else:
        print(DOC_STRING)
        exit()
